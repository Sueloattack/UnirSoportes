import html
from copy import copy
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from PySide6.QtCore import QObject, Signal

from buscar_soportes_nu import (
    ejecutar_busqueda_soportes_nu,
    ejecutar_compresion_en_lotes,
    ejecutar_validacion_soportes,
)


COLOR_INFO = "#5DADE2"
COLOR_SUCCESS = "#2ECC71"
COLOR_WARNING = "#F39C12"
COLOR_ERROR = "#E74C3C"
COLOR_DEFAULT = "#ECF0F1"


def _normalizar_columnas(columnas):
    return tuple(str(columna).strip() for columna in columnas)


def _copiar_estilo_celda(origen, destino):
    if origen.has_style:
        destino.font = copy(origen.font)
        destino.fill = copy(origen.fill)
        destino.border = copy(origen.border)
        destino.alignment = copy(origen.alignment)
        destino.number_format = origen.number_format
        destino.protection = copy(origen.protection)


def _guardar_consolidado_con_formato(template_path: Path, nombre_hoja: str, dataframe: pd.DataFrame, ruta_salida: Path):
    workbook = load_workbook(template_path)
    hoja = workbook[nombre_hoja]

    celdas_referencia = {}
    alturas_referencia = {}
    if hoja.max_row >= 2:
        for columna in range(1, hoja.max_column + 1):
            celdas_referencia[columna] = copy(hoja.cell(row=2, column=columna))
        for fila in range(2, hoja.max_row + 1):
            if hoja.row_dimensions[fila].height is not None:
                alturas_referencia[fila] = hoja.row_dimensions[fila].height

    if hoja.max_row > 1:
        hoja.delete_rows(2, hoja.max_row - 1)

    for fila in dataframe.itertuples(index=False, name=None):
        hoja.append(list(fila))

    if celdas_referencia:
        altura_base = alturas_referencia.get(2)
        for fila_idx in range(2, hoja.max_row + 1):
            if altura_base is not None:
                hoja.row_dimensions[fila_idx].height = altura_base
            for columna_idx in range(1, hoja.max_column + 1):
                _copiar_estilo_celda(celdas_referencia[columna_idx], hoja.cell(row=fila_idx, column=columna_idx))

    workbook.save(ruta_salida)


def consolidar_archivos_excel(
    carpeta_origen: str,
    carpeta_destino: str,
    nombre_salida: str,
    incluir_subcarpetas: bool = False,
    logger=None,
    cancel_check=None,
):
    carpeta_base = Path(carpeta_origen)
    if not carpeta_base.is_dir():
        raise ValueError(f"La ruta '{carpeta_origen}' no es un directorio válido.")

    carpeta_salida = Path(carpeta_destino)
    carpeta_salida.mkdir(parents=True, exist_ok=True)

    nombre_archivo = (nombre_salida or "consolidado_excel.xlsx").strip()
    if not nombre_archivo.lower().endswith(".xlsx"):
        nombre_archivo += ".xlsx"

    def emitir(mensaje, nivel="info"):
        if logger:
            logger(mensaje, nivel)

    patron = "**/*" if incluir_subcarpetas else "*"
    archivos_excel = sorted(
        [
            ruta for ruta in carpeta_base.glob(patron)
            if ruta.is_file() and ruta.suffix.lower() in {".xlsx", ".xls", ".xlsm"}
        ]
    )

    if not archivos_excel:
        raise ValueError("No se encontraron archivos Excel en la carpeta seleccionada.")

    emitir(f"Se encontraron {len(archivos_excel)} archivos Excel para analizar.", "info")

    grupos = {}
    omitidos = []

    for archivo in archivos_excel:
        if cancel_check and cancel_check():
            return {'estado': 'cancelado'}

        try:
            libro = pd.ExcelFile(archivo)
            if not libro.sheet_names:
                omitidos.append({'archivo': str(archivo), 'motivo': 'Sin hojas disponibles'})
                continue

            hoja = libro.sheet_names[0]
            cabecera = pd.read_excel(archivo, sheet_name=hoja, nrows=0)
            columnas = _normalizar_columnas(cabecera.columns)
            if not columnas:
                omitidos.append({'archivo': str(archivo), 'motivo': 'Sin columnas detectables'})
                continue

            firma = (hoja.strip().lower(), columnas)
            grupos.setdefault(firma, []).append(archivo)
        except ImportError as error:
            omitidos.append({'archivo': str(archivo), 'motivo': f'Falta dependencia para leerlo: {error}'})
        except Exception as error:
            omitidos.append({'archivo': str(archivo), 'motivo': str(error)})

    if not grupos:
        raise ValueError("No se pudieron leer archivos Excel compatibles para consolidar.")

    firma_objetivo, archivos_compatibles = max(
        grupos.items(),
        key=lambda item: (len(item[1]), len(item[0][1])),
    )

    if len(archivos_compatibles) < 2:
        raise ValueError("No se encontraron al menos 2 archivos con la misma estructura para consolidar.")

    hoja_objetivo, columnas_objetivo = firma_objetivo
    emitir(
        f"Se consolidarán {len(archivos_compatibles)} archivos compatibles con hoja '{hoja_objetivo}' y {len(columnas_objetivo)} columnas.",
        "success",
    )

    for firma, archivos in grupos.items():
        if firma != firma_objetivo:
            descripcion = f"Estructura distinta: hoja '{firma[0]}' con {len(firma[1])} columnas"
            for archivo in archivos:
                omitidos.append({'archivo': str(archivo), 'motivo': descripcion})

    dataframes = []
    archivo_plantilla = archivos_compatibles[0]
    hoja_real = pd.ExcelFile(archivo_plantilla).sheet_names[0]
    for indice, archivo in enumerate(archivos_compatibles, start=1):
        if cancel_check and cancel_check():
            return {'estado': 'cancelado'}

        emitir(f"Leyendo {indice}/{len(archivos_compatibles)}: {archivo.name}", "info")
        dataframe = pd.read_excel(archivo, sheet_name=hoja_real)
        dataframe.columns = list(_normalizar_columnas(dataframe.columns))
        dataframes.append(dataframe)

    consolidado = pd.concat(dataframes, ignore_index=True)
    ruta_salida = carpeta_salida / nombre_archivo

    _guardar_consolidado_con_formato(archivo_plantilla, hoja_real, consolidado, ruta_salida)

    emitir(f"Excel consolidado creado en: {ruta_salida}", "success")
    if omitidos:
        emitir(f"Se omitieron {len(omitidos)} archivos por estructura distinta o error de lectura.", "warning")

    return {
        'estado': 'completado',
        'archivo_salida': str(ruta_salida),
        'archivos_encontrados': len(archivos_excel),
        'archivos_consolidados': len(archivos_compatibles),
        'filas_consolidadas': len(consolidado),
        'omitidos': len(omitidos),
    }


class FuncionalidadesPreviWorker(QObject):
    progreso_actualizado = Signal(str)
    proceso_finalizado = Signal(dict)

    def __init__(self, parametros, modo):
        super().__init__()
        self.parametros = parametros
        self.modo = modo
        self.esta_cancelado = False

    def cancelar(self):
        self.esta_cancelado = True

    def ejecutar(self):
        try:
            if self.modo == "buscar":
                resultado = ejecutar_busqueda_soportes_nu(
                    self.parametros['modo_busqueda'],
                    self.parametros['facturas'],
                    self.parametros['carpeta_origen'],
                    self.parametros['carpeta_destino'],
                    logger=self._emitir_log,
                    cancel_check=self._debe_cancelarse,
                )
            elif self.modo == "comprimir":
                resultado = ejecutar_compresion_en_lotes(
                    self.parametros['modo_compresion'],
                    self.parametros['carpeta_origen'],
                    self.parametros['carpeta_destino'],
                    logger=self._emitir_log,
                    cancel_check=self._debe_cancelarse,
                )
            elif self.modo == "validar":
                resultado = ejecutar_validacion_soportes(
                    self.parametros['carpeta_validacion'],
                    self.parametros['facturas'],
                    self.parametros['modo_validacion'],
                    logger=self._emitir_log,
                    cancel_check=self._debe_cancelarse,
                )
            elif self.modo == "unir_excels":
                resultado = consolidar_archivos_excel(
                    self.parametros['carpeta_origen'],
                    self.parametros['carpeta_destino'],
                    self.parametros['nombre_salida'],
                    incluir_subcarpetas=self.parametros.get('incluir_subcarpetas', False),
                    logger=self._emitir_log,
                    cancel_check=self._debe_cancelarse,
                )
            else:
                raise ValueError(f"Modo desconocido: {self.modo}")

            self.proceso_finalizado.emit(resultado)
        except Exception as error:
            self.proceso_finalizado.emit({'error': str(error)})

    def _debe_cancelarse(self):
        return self.esta_cancelado

    def _emitir_log(self, mensaje: str, nivel: str = "info"):
        color = {
            'info': COLOR_INFO,
            'success': COLOR_SUCCESS,
            'warning': COLOR_WARNING,
            'error': COLOR_ERROR,
        }.get(nivel, COLOR_DEFAULT)
        contenido = html.escape(str(mensaje)).replace("\n", "<br>")
        self.progreso_actualizado.emit(f"<p style='color:{color}; margin:0;'>{contenido}</p>")