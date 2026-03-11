from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                                QLineEdit, QGroupBox, QLabel, QDateEdit, QCheckBox,
                                QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar,
                                QMessageBox, QFileDialog)
from PySide6.QtCore import QThread, Qt, Signal, QDate
from PySide6.QtGui import QColor
from gui.common.componentes_comunes import crear_selector_carpeta
import os
import re
from typing import Optional

class AutomatizadorRadicacionWorker(QThread):
    """Worker thread para ejecutar la automatización en segundo plano"""
    progreso_actualizado = Signal(str, int)
    proceso_finalizado = Signal(dict)
    error_ocurrido = Signal(str)
    
    def __init__(self, carpeta_pdfs, fecha_notificacion, username, password, headless):
        super().__init__()
        self.carpeta_pdfs = carpeta_pdfs
        self.fecha_notificacion = fecha_notificacion
        self.username = username
        self.password = password
        self.headless = headless
        self._cancelado = False
    
    def run(self):
        print("\n" + "="*70)
        print("WORKER THREAD INICIADO")
        try:
            from logica.workers.automatizador_radicacion_logic import automatizar_radicacion_logic
            
            automatizar_radicacion_logic(
                self.carpeta_pdfs,
                self.fecha_notificacion,
                self.username,
                self.password,
                self.headless,
                self.progreso_actualizado,
                self.proceso_finalizado,
                self.error_ocurrido
            )
        except Exception as e:
            print(f"ERROR CRÍTICO en worker: {e}")
            import traceback
            traceback.print_exc()
            self.error_ocurrido.emit(f"Error crítico: {str(e)}")
    
    def cancelar(self):
        self._cancelado = True

class AutomatizadorRadicacionWidget(QWidget):
    """Widget para automatizar la radicación de cartas glosas"""
    
    def __init__(self, parent=None):
        super(AutomatizadorRadicacionWidget, self).__init__(parent)
        self.worker: Optional[AutomatizadorRadicacionWorker] = None
        self._ultimos_resultados: dict = {}

        # Colores (texto oscuro para contraste)
        self.color_exito      = QColor(39, 174,  96)   # Verde
        self.color_error      = QColor(192,  57,  43)  # Rojo
        self.color_advertencia= QColor(211, 142,   0)  # Amarillo oscuro
        self.color_texto_filas= QColor(255, 255, 255)  # Blanco

        self.init_ui()
    
    def init_ui(self):
        layout_principal = QVBoxLayout(self)
        layout_principal.setContentsMargins(20, 20, 20, 20)
        layout_principal.setSpacing(15)
        
        # --- Título ---
        label_titulo = QLabel("Automatizador de Radicación Previsora")
        label_titulo.setObjectName("AyudaTitulo")
        label_titulo.setAlignment(Qt.AlignCenter)
        layout_principal.addWidget(label_titulo)
        
        # --- Configuración ---
        grupo_config = QGroupBox("Configuración")
        layout_config = QVBoxLayout()
        
        self.ruta_carpeta_line_edit, selector_carpeta_layout = crear_selector_carpeta(
            "Carpeta con PDFs:", "Seleccionar Carpeta", self.validar_formulario
        )
        layout_config.addLayout(selector_carpeta_layout)
        
        # Fecha
        layout_fecha = QHBoxLayout()
        label_fecha = QLabel("Fecha de Notificación:")
        label_fecha.setMinimumWidth(150)
        self.fecha_edit = QDateEdit()
        self.fecha_edit.setCalendarPopup(True)
        self.fecha_edit.setDate(QDate.currentDate())
        self.fecha_edit.setMaximumDate(QDate.currentDate())
        self.fecha_edit.setDisplayFormat("dd/MM/yyyy")
        layout_fecha.addWidget(label_fecha)
        layout_fecha.addWidget(self.fecha_edit)
        layout_fecha.addStretch()
        layout_config.addLayout(layout_fecha)
        
        # Credenciales
        layout_usuario = QHBoxLayout()
        label_usuario = QLabel("Usuario:")
        label_usuario.setMinimumWidth(150)
        self.usuario_edit = QLineEdit()
        self.usuario_edit.setText("AJ.JOSE") # Default
        layout_usuario.addWidget(label_usuario)
        layout_usuario.addWidget(self.usuario_edit)
        layout_config.addLayout(layout_usuario)
        
        layout_pass = QHBoxLayout()
        label_pass = QLabel("Contraseña:")
        label_pass.setMinimumWidth(150)
        self.password_edit = QLineEdit()
        self.password_edit.setText("1005911366") # Default
        self.password_edit.setEchoMode(QLineEdit.Password)
        layout_pass.addWidget(label_pass)
        layout_pass.addWidget(self.password_edit)
        layout_config.addLayout(layout_pass)
        
        self.headless_checkbox = QCheckBox("Modo invisible (Headless)")
        self.headless_checkbox.setChecked(False)
        layout_config.addWidget(self.headless_checkbox)
        
        grupo_config.setLayout(layout_config)
        layout_principal.addWidget(grupo_config)
        
        # --- Botones ---
        layout_botones = QHBoxLayout()
        self.boton_iniciar = QPushButton("Iniciar Radicación")
        self.boton_iniciar.setObjectName("BotonPrimario")
        self.boton_iniciar.clicked.connect(self.iniciar_automatizacion)
        self.boton_iniciar.setEnabled(False)

        self.boton_cancelar = QPushButton("Detener")
        self.boton_cancelar.setObjectName("BotonSecundario")
        self.boton_cancelar.clicked.connect(self.cancelar_automatizacion)
        self.boton_cancelar.setEnabled(False)

        self.boton_exportar = QPushButton("📊 Exportar Resultados")
        self.boton_exportar.setObjectName("BotonSecundario")
        self.boton_exportar.clicked.connect(self.exportar_excel)
        self.boton_exportar.setEnabled(False)

        layout_botones.addWidget(self.boton_iniciar)
        layout_botones.addWidget(self.boton_cancelar)
        layout_botones.addWidget(self.boton_exportar)
        layout_principal.addLayout(layout_botones)
        
        # --- Progreso ---
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        layout_principal.addWidget(self.progress_bar)
        
        self.label_estado = QLabel("Esperando configuración...")
        self.label_estado.setAlignment(Qt.AlignCenter)
        layout_principal.addWidget(self.label_estado)
        
        # --- Resultados ---
        self.tabla_resultados = QTableWidget()
        self.tabla_resultados.setColumnCount(4)
        self.tabla_resultados.setHorizontalHeaderLabels(["Factura", "Estado", "Valor", "Tipo"])
        self.tabla_resultados.setShowGrid(True)
        self.tabla_resultados.setAlternatingRowColors(False)
        self.tabla_resultados.verticalHeader().setVisible(False)  # quita la columna blanca de índice

        header = self.tabla_resultados.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)

        layout_principal.addWidget(self.tabla_resultados)
    
    def validar_formulario(self):
        carpeta = self.ruta_carpeta_line_edit.text()
        ok = bool(carpeta and os.path.isdir(carpeta) and 
                  any(f.lower().endswith('.pdf') for f in os.listdir(carpeta)))
        self.boton_iniciar.setEnabled(ok)
        self.label_estado.setText("✓ Carpeta lista" if ok else "Seleccione una carpeta válida")

    def iniciar_automatizacion(self):
        carpeta = self.ruta_carpeta_line_edit.text()
        fecha_str = self.fecha_edit.date().toString("dd/MM/yyyy")
        user = self.usuario_edit.text()
        pwd = self.password_edit.text()
        headless = self.headless_checkbox.isChecked()
        
        # Reset UI
        self.tabla_resultados.setRowCount(0)
        self.progress_bar.setValue(0)
        self.set_controles_habilitados(False)
        
        worker = AutomatizadorRadicacionWorker(carpeta, fecha_str, user, pwd, headless)
        self.worker = worker
        if worker is not None:
            worker.progreso_actualizado.connect(self.actualizar_progreso)
            worker.proceso_finalizado.connect(self.finalizar_proceso)
            worker.error_ocurrido.connect(self.mostrar_error)
            worker.start()
        
        self.label_estado.setText("Iniciando motor de radicación...")

    def actualizar_progreso(self, msj, val):
        self.progress_bar.setValue(val)
        self.label_estado.setText(msj)

    def finalizar_proceso(self, resultados):
        """Maneja el fin del proceso: llena tabla y muestra popup"""
        self._ultimos_resultados = resultados

        exitosos    = resultados.get('exitosos', [])
        fallidos    = resultados.get('fallidos', [])
        advertencias= resultados.get('advertencias', [])

        count_ok   = len(exitosos)
        count_err  = len(fallidos)
        count_warn = len(advertencias)
        total      = count_ok + count_err + count_warn

        for item in exitosos:     self._agregar_fila_tabla(item, "OK")
        for item in advertencias: self._agregar_fila_tabla(item, "WARN")
        for item in fallidos:     self._agregar_fila_tabla(item, "ERR")

        self.progress_bar.setValue(100)
        ok_txt   = f"✅ {count_ok} radicados" if count_ok   else ""
        warn_txt = f"⚠️ {count_warn} saltados" if count_warn else ""
        err_txt  = f"❌ {count_err} errores"   if count_err  else ""
        partes   = [p for p in [ok_txt, warn_txt, err_txt] if p]
        self.label_estado.setText("  |  ".join(partes) if partes else "Sin resultados.")

        self.set_controles_habilitados(True)
        self.boton_exportar.setEnabled(True)

    def _hacer_item(self, texto, color_fondo):
        """Crea un QTableWidgetItem coloreado con texto blanco."""
        it = QTableWidgetItem(texto)
        it.setBackground(color_fondo)
        it.setForeground(self.color_texto_filas)
        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
        return it

    def _colorear_fila(self, fila, color):
        """Aplica un color de fondo a todas las celdas de una fila."""
        for col in range(self.tabla_resultados.columnCount()):
            it = self.tabla_resultados.item(fila, col)
            if it:
                it.setBackground(color)
                it.setForeground(self.color_texto_filas)

    def _agregar_fila_tabla(self, item, estado_code):
        fila = self.tabla_resultados.rowCount()
        self.tabla_resultados.insertRow(fila)

        nombre_arch  = item.get('archivo', '')
        match = re.search(r'([A-Za-z]+)(\d+)', nombre_arch)
        factura_str = f"{match.group(1).upper()}{match.group(2)}" if match else nombre_arch.upper().replace('.PDF', '')

        if estado_code == "OK":
            color      = self.color_exito
            valor      = item.get('valor', 0) or 0
            tipo_col   = item.get('tipo', 'GP')
            estado_txt = "EXITOSO"
            valor_txt  = f"$ {valor:,.0f}"
        elif estado_code == "WARN":
            color      = self.color_advertencia
            estado_txt = "SALTADO"
            valor_txt  = "-"
            tipo_col   = "N/A"
        else:
            color      = self.color_error
            estado_txt = "ERROR"
            valor_txt  = item.get('error', 'Desconocido')[:30]
            tipo_col   = "-"

        self.tabla_resultados.setItem(fila, 0, self._hacer_item(factura_str, color))
        self.tabla_resultados.setItem(fila, 1, self._hacer_item(estado_txt,  color))
        self.tabla_resultados.setItem(fila, 2, self._hacer_item(valor_txt,   color))
        self.tabla_resultados.setItem(fila, 3, self._hacer_item(tipo_col,    color))

    def mostrar_error(self, msj):
        self.label_estado.setText(f"❌ {msj}")
        QMessageBox.critical(self, "Error de Ejecución", f"Ocurrió un error:\n{msj}")
        self.set_controles_habilitados(True)
        self.boton_exportar.setEnabled(bool(self._ultimos_resultados))

    def cancelar_automatizacion(self):
        worker = self.worker
        if worker is not None and worker.isRunning():
            worker.cancelar()
            worker.terminate()
            self.label_estado.setText("Cancelado por usuario.")
            self.set_controles_habilitados(True)
            self.boton_exportar.setEnabled(bool(self._ultimos_resultados))

    def exportar_excel(self):
        """Exporta los resultados a un archivo Excel."""
        try:                                           # Bug 4
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            QMessageBox.critical(self, "Dependencia faltante",
                                 "openpyxl no está instalado.\n\nEjecuta:\n  pip install openpyxl")
            return

        res  = self._ultimos_resultados
        if not res:
            QMessageBox.warning(self, "Sin datos", "No hay resultados para exportar.")
            return

        ruta, _ = QFileDialog.getSaveFileName(
            self, "Guardar reporte", "resultados_radicacion.xlsx",
            "Excel (*.xlsx)"
        )
        if not ruta:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resultados"

        # Estilos
        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_font  = Font(bold=True, color="FFFFFF")
        ok_fill   = PatternFill("solid", fgColor="1E8449")
        err_fill  = PatternFill("solid", fgColor="922B21")
        warn_fill = PatternFill("solid", fgColor="9A7D0A")
        blanco    = Font(color="FFFFFF", bold=True)
        num_fmt   = '"$" #,##0'   # formato numérico con $

        # Encabezados — 4 columnas
        ws.append(["Factura", "Estado", "Valor", "Tipo"])
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

        def _agregar(items, estado_label, fill_color):
            for it in items:
                nombre_arch = it.get('archivo', '')
                m = re.search(r'([A-Za-z]+)(\d+)', nombre_arch)
                factura = f"{m.group(1).upper()}{m.group(2)}" if m else nombre_arch.upper().replace('.PDF','')

                if estado_label == "EXITOSO":
                    valor_num = it.get('valor', 0) or 0
                    tipo_txt  = it.get('tipo', 'GP')
                else:
                    valor_num = None
                    tipo_txt  = it.get('razon', '') or it.get('error', '')

                row_idx = ws.max_row + 1
                ws.append([factura, estado_label, valor_num, tipo_txt])
                for col in range(1, 5):
                    c = ws.cell(row=row_idx, column=col)
                    c.fill = fill_color
                    c.font = blanco
                    c.alignment = Alignment(horizontal="center")
                # Columna Valor: alinear derecha y formato $
                vc = ws.cell(row=row_idx, column=3)
                vc.alignment = Alignment(horizontal="right")
                if valor_num is not None:
                    vc.number_format = num_fmt

        _agregar(res.get('exitosos', []),     "EXITOSO",  ok_fill)
        _agregar(res.get('advertencias', []), "SALTADO",  warn_fill)
        _agregar(res.get('fallidos', []),     "ERROR",    err_fill)

        # Anchos de columna
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 10

        wb.save(ruta)
        QMessageBox.information(self, "Exportado", f"Reporte guardado en:\n{ruta}")

    def set_controles_habilitados(self, enable):
        self.ruta_carpeta_line_edit.setEnabled(enable)
        self.boton_cancelar.setEnabled(not enable)
        if enable:
            self.validar_formulario()   # Bug 3: re-evalúa si la carpeta sigue siendo válida
        else:
            self.boton_iniciar.setEnabled(False)